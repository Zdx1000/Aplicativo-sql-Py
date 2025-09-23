from __future__ import annotations

from time import time
from typing import Optional
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtGui import QIntValidator
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QFormLayout,
    QLineEdit,
    QComboBox,
    QPushButton,
    QLabel,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractScrollArea,
    QDialog,
)

from style import QSS_CONSULTAS_PAGE
try:
    from config import SETORES as SETORES_GLOBAIS
except Exception:
    SETORES_GLOBAIS = None


# Utilitário simples para ícone (opcional)
def _resource_path(rel_path: str) -> str:
    base = getattr(__import__("sys"), "_MEIPASS", None)
    if base:
        return str(Path(base) / rel_path)
    return str(Path(__file__).resolve().parent / rel_path)


def _get_app_icon():
    from PySide6.QtGui import QIcon
    for c in ["assets/app_icon.ico", "assets/app_icon.png", "assets/app_icon.svg"]:
        p = Path(_resource_path(c))
        if p.exists():
            return QIcon(str(p))
    return QIcon()


class ConsultaExportDialog(QDialog):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Exportar Resultados")
        self.setWindowIcon(_get_app_icon())
        self._build()
        self._formato = "csv"

    def _build(self) -> None:
        from PySide6.QtWidgets import QVBoxLayout, QLabel, QComboBox, QHBoxLayout
        lay = QVBoxLayout(self)
        lab = QLabel("Escolha o formato para exportar os resultados atuais:")
        lab.setWordWrap(True)
        lay.addWidget(lab)
        self.cb_formato = QComboBox()
        self.cb_formato.addItems(["csv", "xlsx", "txt"])
        lay.addWidget(self.cb_formato)
        row = QHBoxLayout()
        btn_cancel = QPushButton("Cancelar")
        btn_ok = QPushButton("Exportar")
        btn_cancel.clicked.connect(self.reject)
        btn_ok.clicked.connect(self._on_ok)
        row.addStretch(1)
        row.addWidget(btn_cancel)
        row.addWidget(btn_ok)
        lay.addLayout(row)
        self.resize(300, 150)

    def _on_ok(self) -> None:
        self._formato = self.cb_formato.currentText()
        self.accept()

    def get_formato(self) -> str:
        return self._formato


class ConsultasPage(QWidget):
    """Página de Consultas (Bloqueado, Consolidado e Reimpressão)."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("PaginaConsultas")
        # cache simples p/ item
        self._consulta_cache: dict[int, tuple[float, list[dict]]] = {}
        self._consulta_cache_ttl = 120.0
        self._consulta_cache_max = 50
        self._ultimos_resultados_consulta: list[dict] = []
        self._build()

    def _build(self) -> None:
        page = self
        lay = QVBoxLayout(page)
        lay.setContentsMargins(32, 24, 32, 24)
        lay.setSpacing(16)

        form = QFormLayout()
        self.form_consultas = form
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        self.cb_fonte = QComboBox()
        self.cb_fonte.addItems(["Bloqueado", "Consolidado", "Reimpressão", "Almoxarifado", "EPIs", "Senha Corte"])
        self.cb_fonte.currentTextChanged.connect(self._on_mudar_fonte_consulta)
        form.addRow("Fonte:", self.cb_fonte)

        # Bloqueado
        self.ed_busca_item = QLineEdit()
        self.ed_busca_item.setPlaceholderText("Digite código do item ou % para todos e Enter...")
        self.ed_busca_item.returnPressed.connect(self._executar_consulta_item)
        self.btn_buscar_item = QPushButton("Buscar")
        self.btn_buscar_item.clicked.connect(self._executar_consulta_item)
        self.btn_exportar_item = QPushButton("Exportar")
        self.btn_exportar_item.setToolTip("Exportar resultados atuais")
        self.btn_exportar_item.clicked.connect(self._exportar_consulta)
        linha_busca = QHBoxLayout()
        linha_busca.addWidget(self.ed_busca_item, 1)
        linha_busca.addWidget(self.btn_buscar_item)
        linha_busca.addWidget(self.btn_exportar_item)
        self.wrap_busca = QWidget()
        self.wrap_busca.setLayout(linha_busca)
        form.addRow("Item:", self.wrap_busca)

        # Consolidado (datas)
        from PySide6.QtWidgets import QDateEdit
        from PySide6.QtCore import QDate
        self.ed_data_ini = QDateEdit()
        self.ed_data_ini.setDisplayFormat("yyyy-MM-dd")
        self.ed_data_ini.setCalendarPopup(True)
        self.ed_data_ini.setDate(QDate.currentDate())
        self.ed_data_fim = QDateEdit()
        self.ed_data_fim.setDisplayFormat("yyyy-MM-dd")
        self.ed_data_fim.setCalendarPopup(True)
        self.ed_data_fim.setDate(QDate.currentDate())
        self.btn_buscar_periodo = QPushButton("Buscar")
        self.btn_buscar_periodo.clicked.connect(self._executar_consulta_consolidado)
        self.btn_exportar_periodo = QPushButton("Exportar")
        self.btn_exportar_periodo.setToolTip("Exportar resultados atuais")
        self.btn_exportar_periodo.clicked.connect(self._exportar_consulta)
        row_dt = QHBoxLayout()
        row_dt.addWidget(self.ed_data_ini)
        row_dt.addWidget(self.ed_data_fim)
        row_dt.addWidget(self.btn_buscar_periodo)
        row_dt.addWidget(self.btn_exportar_periodo)
        self.wrap_dt = QWidget()
        self.wrap_dt.setLayout(row_dt)
        form.addRow("Período:", self.wrap_dt)

        # Reimpressão (Monitoramento)
        self.cb_campo_mon = QComboBox()
        self.cb_campo_mon.addItems(["Onda", "Carga", "Container"])
        self.ed_busca_mon = QLineEdit()
        self.ed_busca_mon.setPlaceholderText("Digite valor ou % para todos e Enter...")
        self.ed_busca_mon.returnPressed.connect(self._executar_consulta_monitoramento)
        self.btn_buscar_mon = QPushButton("Buscar")
        self.btn_buscar_mon.clicked.connect(self._executar_consulta_monitoramento)
        self.btn_exportar_mon = QPushButton("Exportar")
        self.btn_exportar_mon.setToolTip("Exportar resultados atuais")
        self.btn_exportar_mon.clicked.connect(self._exportar_consulta)
        row_mon = QHBoxLayout()
        row_mon.addWidget(self.cb_campo_mon)
        row_mon.addWidget(self.ed_busca_mon, 1)
        row_mon.addWidget(self.btn_buscar_mon)
        row_mon.addWidget(self.btn_exportar_mon)
        self.wrap_mon = QWidget()
        self.wrap_mon.setLayout(row_mon)
        form.addRow("Filtro:", self.wrap_mon)

        # Almoxarifado (Turno + datas)
        from PySide6.QtWidgets import QDateEdit
        from PySide6.QtCore import QDate
        self.cb_turno_alm = QComboBox()
        self.cb_turno_alm.addItems(["Todos", "1° Turno", "2° Turno"])
        self.ed_alm_data_ini = QDateEdit()
        self.ed_alm_data_ini.setDisplayFormat("yyyy-MM-dd")
        self.ed_alm_data_ini.setCalendarPopup(True)
        self.ed_alm_data_ini.setDate(QDate.currentDate())
        self.ed_alm_data_fim = QDateEdit()
        self.ed_alm_data_fim.setDisplayFormat("yyyy-MM-dd")
        self.ed_alm_data_fim.setCalendarPopup(True)
        self.ed_alm_data_fim.setDate(QDate.currentDate())
        self.btn_buscar_alm = QPushButton("Buscar")
        self.btn_buscar_alm.clicked.connect(self._executar_consulta_almox)
        self.btn_exportar_alm = QPushButton("Exportar")
        self.btn_exportar_alm.setToolTip("Exportar resultados atuais")
        self.btn_exportar_alm.clicked.connect(self._exportar_consulta)
        row_alm = QHBoxLayout()
        row_alm.addWidget(self.cb_turno_alm)
        row_alm.addWidget(self.ed_alm_data_ini)
        row_alm.addWidget(self.ed_alm_data_fim)
        row_alm.addWidget(self.btn_buscar_alm)
        row_alm.addWidget(self.btn_exportar_alm)
        self.wrap_alm = QWidget()
        self.wrap_alm.setLayout(row_alm)
        form.addRow("Filtro:", self.wrap_alm)

        # EPIs (por período OU por matrícula)
        from PySide6.QtWidgets import QDateEdit
        from PySide6.QtCore import QDate
        self.ed_epi_data_ini = QDateEdit()
        self.ed_epi_data_ini.setDisplayFormat("yyyy-MM-dd")
        self.ed_epi_data_ini.setCalendarPopup(True)
        self.ed_epi_data_ini.setDate(QDate.currentDate())
        self.ed_epi_data_fim = QDateEdit()
        self.ed_epi_data_fim.setDisplayFormat("yyyy-MM-dd")
        self.ed_epi_data_fim.setCalendarPopup(True)
        self.ed_epi_data_fim.setDate(QDate.currentDate())
        self.ed_epi_matricula = QLineEdit()
        self.ed_epi_matricula.setPlaceholderText("Matrícula (opcional)")
        try:
            self.ed_epi_matricula.setValidator(QIntValidator(1, 99_999_999, self))
        except Exception:
            pass
        self.btn_buscar_epis = QPushButton("Buscar")
        self.btn_buscar_epis.clicked.connect(self._executar_consulta_epis)
        self.btn_exportar_epis = QPushButton("Exportar")
        self.btn_exportar_epis.setToolTip("Exportar resultados atuais")
        self.btn_exportar_epis.clicked.connect(self._exportar_consulta)
        row_epi = QHBoxLayout()
        row_epi.addWidget(QLabel("De:"))
        row_epi.addWidget(self.ed_epi_data_ini)
        row_epi.addWidget(QLabel("Até:"))
        row_epi.addWidget(self.ed_epi_data_fim)
        row_epi.addWidget(QLabel("Matrícula:"))
        row_epi.addWidget(self.ed_epi_matricula)
        row_epi.addWidget(self.btn_buscar_epis)
        row_epi.addWidget(self.btn_exportar_epis)
        self.wrap_epi = QWidget()
        self.wrap_epi.setLayout(row_epi)
        form.addRow("Filtro:", self.wrap_epi)

        # Senha Corte (por período e ordem)
        from PySide6.QtWidgets import QDateEdit
        from PySide6.QtCore import QDate
        self.ed_sc_ordem = QLineEdit()
        self.ed_sc_ordem.setValidator(QIntValidator(10000, 99999999, self))
        self.ed_sc_ordem.setPlaceholderText("Ordem (opcional)")
        self.ed_sc_data_ini = QDateEdit()
        self.ed_sc_data_ini.setDisplayFormat("yyyy-MM-dd")
        self.ed_sc_data_ini.setCalendarPopup(True)
        self.ed_sc_data_ini.setDate(QDate.currentDate())
        self.ed_sc_data_fim = QDateEdit()
        self.ed_sc_data_fim.setDisplayFormat("yyyy-MM-dd")
        self.ed_sc_data_fim.setCalendarPopup(True)
        self.ed_sc_data_fim.setDate(QDate.currentDate())
        self.btn_buscar_sc = QPushButton("Buscar")
        self.btn_buscar_sc.clicked.connect(self._executar_consulta_senha_corte)
        self.btn_exportar_sc = QPushButton("Exportar")
        self.btn_exportar_sc.setToolTip("Exportar resultados atuais")
        self.btn_exportar_sc.clicked.connect(self._exportar_consulta)
        row_sc = QHBoxLayout()
        row_sc.addWidget(QLabel("Ordem:"))
        row_sc.addWidget(self.ed_sc_ordem)
        row_sc.addWidget(QLabel("De:"))
        row_sc.addWidget(self.ed_sc_data_ini)
        row_sc.addWidget(QLabel("Até:"))
        row_sc.addWidget(self.ed_sc_data_fim)
        row_sc.addWidget(self.btn_buscar_sc)
        row_sc.addWidget(self.btn_exportar_sc)
        self.wrap_sc = QWidget()
        self.wrap_sc.setLayout(row_sc)
        form.addRow("Filtro:", self.wrap_sc)

        lay.addLayout(form)

        # Tabela resultados (default com 9 colunas de Bloqueado)
        self.tab_resultados = QTableWidget(0, 9)
        self.tab_resultados.setObjectName("TabelaConsultas")
        self.tab_resultados.setHorizontalHeaderLabels(["ID", "Item", "Quantidade", "Motivo", "Setor", "Usuário", "Criado em", "Editar", "Excluir"])
        hdr = self.tab_resultados.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        hdr.setStretchLastSection(True)
        self.tab_resultados.setSizeAdjustPolicy(QAbstractScrollArea.SizeAdjustPolicy.AdjustToContents)
        self.tab_resultados.setAlternatingRowColors(True)
        self.tab_resultados.setSortingEnabled(True)
        lay.addWidget(self.tab_resultados, 1)

        self.lab_status_consulta = QLabel("Informe um item ou período para consultar.")
        self.lab_status_consulta.setObjectName("StatusConsultaLabel")
        lay.addWidget(self.lab_status_consulta)

        page.setStyleSheet(QSS_CONSULTAS_PAGE)
        self._on_mudar_fonte_consulta(self.cb_fonte.currentText())

    # -------- Handlers -------- #
    def _executar_consulta_item(self) -> None:
        texto = self.ed_busca_item.text().strip()
        if not texto:
            self.lab_status_consulta.setText("Digite um código de item ou % para todos.")
            return
        if texto == "%":
            from database import consultar_todos_registros
            regs = consultar_todos_registros()
            self._popular_tabela_consultas(regs)
            self._ultimos_resultados_consulta = regs
            self.lab_status_consulta.setText(f"{len(regs)} registro(s) (todos) exibidos." if regs else "Nenhum registro encontrado.")
            return
        try:
            item_cod = int(texto)
        except ValueError:
            self.lab_status_consulta.setText("Código inválido (use número ou %).")
            return
        now_ts = time()
        cache_hit = False
        entry = self._consulta_cache.get(item_cod)
        if entry and (now_ts - entry[0]) < self._consulta_cache_ttl:
            regs = entry[1]
            cache_hit = True
        else:
            from database import consultar_registros_por_item
            try:
                regs = consultar_registros_por_item(item_cod)
            except ValueError as exc:
                self.lab_status_consulta.setText(str(exc))
                return
            self._consulta_cache[item_cod] = (now_ts, regs)
            if len(self._consulta_cache) > self._consulta_cache_max:
                mais_antigo = min(self._consulta_cache.items(), key=lambda kv: kv[1][0])[0]
                self._consulta_cache.pop(mais_antigo, None)
        self._popular_tabela_consultas(regs)
        self._ultimos_resultados_consulta = regs
        if regs:
            prefixo = "(cache) " if cache_hit else ""
            self.lab_status_consulta.setText(f"{prefixo}{len(regs)} registro(s) encontrados para item {item_cod}.")
        else:
            self.lab_status_consulta.setText("Nenhum registro encontrado.")

    def _executar_consulta_consolidado(self) -> None:
        ini = self.ed_data_ini.date().toString("yyyy-MM-dd")
        fim = self.ed_data_fim.date().toString("yyyy-MM-dd")
        try:
            from database import consultar_consolidado_por_periodo
            regs = consultar_consolidado_por_periodo(data_ini=ini, data_fim=fim)
        except ValueError as exc:
            self.lab_status_consulta.setText(str(exc))
            return
        self._popular_tabela_consolidado(regs)
        self._ultimos_resultados_consulta = regs
        self.lab_status_consulta.setText(f"{len(regs)} linha(s) do Consolidado no período." if regs else "Nenhum registro do Consolidado no período.")

    def _executar_consulta_monitoramento(self) -> None:
        campo_vis = self.cb_campo_mon.currentText() if hasattr(self, "cb_campo_mon") else "Onda"
        valor = self.ed_busca_mon.text().strip() if hasattr(self, "ed_busca_mon") else ""
        if not valor:
            self.lab_status_consulta.setText("Digite um valor (ou % para todos).")
            return
        campo = campo_vis.strip().lower()
        try:
            from database import consultar_monitoramento_por_campo
            regs = consultar_monitoramento_por_campo(campo=campo, valor=valor)
        except ValueError as exc:
            self.lab_status_consulta.setText(str(exc))
            return
        except Exception as exc:
            self.lab_status_consulta.setText(f"Erro na consulta: {exc}")
            return
        self._popular_tabela_monitoramento(regs)
        self._ultimos_resultados_consulta = regs
        self.lab_status_consulta.setText(f"{len(regs)} registro(s) encontrados em Reimpressão." if regs else "Nenhum registro encontrado.")

    def _on_mudar_fonte_consulta(self, fonte: str) -> None:
        is_bloq = (fonte == "Bloqueado")
        is_consol = (fonte == "Consolidado")
        is_mon = (fonte == "Reimpressão")
        is_alm = (fonte == "Almoxarifado")
        is_epi = (fonte == "EPIs")
        is_sc = (fonte == "Senha Corte")
        # Mostrar/ocultar
        self.wrap_busca.setVisible(is_bloq)
        lbl_item = self.form_consultas.labelForField(self.wrap_busca)
        if lbl_item:
            lbl_item.setVisible(is_bloq)
        self.wrap_dt.setVisible(is_consol)
        lbl_dt = self.form_consultas.labelForField(self.wrap_dt)
        if lbl_dt:
            lbl_dt.setVisible(is_consol)
        self.wrap_mon.setVisible(is_mon)
        lbl_mon = self.form_consultas.labelForField(self.wrap_mon)
        if lbl_mon:
            lbl_mon.setVisible(is_mon)
        self.wrap_alm.setVisible(is_alm)
        lbl_alm = self.form_consultas.labelForField(self.wrap_alm)
        if lbl_alm:
            lbl_alm.setVisible(is_alm)
        self.wrap_epi.setVisible(is_epi)
        lbl_epi = self.form_consultas.labelForField(self.wrap_epi)
        if lbl_epi:
            lbl_epi.setVisible(is_epi)
        self.wrap_sc.setVisible(is_sc)
        lbl_sc = self.form_consultas.labelForField(self.wrap_sc)
        if lbl_sc:
            lbl_sc.setVisible(is_sc)
        # Habilitar/desabilitar
        self.ed_busca_item.setEnabled(is_bloq)
        self.btn_buscar_item.setEnabled(is_bloq)
        self.ed_data_ini.setEnabled(is_consol)
        self.ed_data_fim.setEnabled(is_consol)
        self.btn_buscar_periodo.setEnabled(is_consol)
        self.cb_campo_mon.setEnabled(is_mon)
        self.ed_busca_mon.setEnabled(is_mon)
        self.btn_buscar_mon.setEnabled(is_mon)
        self.cb_turno_alm.setEnabled(is_alm)
        self.ed_alm_data_ini.setEnabled(is_alm)
        self.ed_alm_data_fim.setEnabled(is_alm)
        self.btn_buscar_alm.setEnabled(is_alm)
        self.ed_epi_data_ini.setEnabled(is_epi)
        self.ed_epi_data_fim.setEnabled(is_epi)
        self.ed_epi_matricula.setEnabled(is_epi)
        self.btn_buscar_epis.setEnabled(is_epi)
        self.ed_sc_data_ini.setEnabled(is_sc)
        self.ed_sc_data_fim.setEnabled(is_sc)
        self.btn_buscar_sc.setEnabled(is_sc)
        # Reset da tabela conforme fonte
        self.tab_resultados.setSortingEnabled(False)
        self.tab_resultados.clear()
        if is_bloq:
            # Adiciona coluna de Data da movimentação após Matrícula
            self.tab_resultados.setColumnCount(11)
            self.tab_resultados.setHorizontalHeaderLabels(["ID", "Item", "Quantidade", "Motivo", "Setor", "Matrícula", "Data Mov.", "Usuário", "Criado em", "Editar", "Excluir"])
            self.lab_status_consulta.setText("Informe um item para consultar.")
        elif is_consol:
            headers = [
                "data_ref",
                "arm.",
                "Descrição Filial",
                "R$ Estoque",
                "Qtde Item Mix",
                "Qtde Item com Estoque",
                "Qtde Item sem Estoque",
                "R$ Bloq. Total",
                "R$ Bloq. no ESTOQUE",
                "R$ Bloq. em Negoc.",
                "R$ Bloq. SALDO",
                "% Item com Estoque",
            ]
            self.tab_resultados.setColumnCount(len(headers))
            self.tab_resultados.setHorizontalHeaderLabels(headers)
            self.lab_status_consulta.setText("Informe um período para consultar.")
        elif is_mon:
            headers = ["ID", "Onda", "Carga", "Container", "Responsável", "Setor", "Observação", "Usuário", "Criado em", "Editar", "Excluir"]
            self.tab_resultados.setColumnCount(len(headers))
            self.tab_resultados.setHorizontalHeaderLabels(headers)
            self.lab_status_consulta.setText("Escolha campo e digite o valor para consultar.")
        elif is_epi:
            headers = ["ID", "Data", "Matrícula", "Setor", "Turno", "Primeira", "Responsável", "Observação", "Usuário", "Criado em", "ITEM", "Excluir"]
            self.tab_resultados.setColumnCount(len(headers))
            self.tab_resultados.setHorizontalHeaderLabels(headers)
            self.lab_status_consulta.setText("Escolha período e/ou matrícula para consultar EPIs.")
        elif is_sc:
            headers = ["ID", "Data Ordem", "Ordem", "Carga", "Valor", "Tipo", "Observação", "Usuário", "Criado em", "ITEM", "Excluir"]
            self.tab_resultados.setColumnCount(len(headers))
            self.tab_resultados.setHorizontalHeaderLabels(headers)
            self.lab_status_consulta.setText("Escolha o período para consultar Senha Corte.")
        else:
            headers = ["ID", "Setor", "Turno", "Matrícula", "Responsável", "Insumo", "Quantidade", "Observação", "Usuário", "Criado em"]
            self.tab_resultados.setColumnCount(len(headers))
            self.tab_resultados.setHorizontalHeaderLabels(headers)
            self.lab_status_consulta.setText("Escolha o turno e período para consultar.")
        self.tab_resultados.setRowCount(0)
        self.tab_resultados.setSortingEnabled(True)

    def _executar_consulta_senha_corte(self) -> None:
        di = self.ed_sc_data_ini.date().toString("yyyy-MM-dd") if hasattr(self, "ed_sc_data_ini") else None
        df = self.ed_sc_data_fim.date().toString("yyyy-MM-dd") if hasattr(self, "ed_sc_data_fim") else None
        ordem_txt = self.ed_sc_ordem.text().strip() if hasattr(self, "ed_sc_ordem") else ""
        try:
            from database import consultar_senha_corte_por_periodo
            if ordem_txt:
                regs = consultar_senha_corte_por_periodo(data_ini=di, data_fim=df)
                regs = [r for r in regs if str(r.get("ordem")) == ordem_txt]
            else:
                regs = consultar_senha_corte_por_periodo(data_ini=di, data_fim=df)
        except ValueError as exc:
            self.lab_status_consulta.setText(str(exc))
            return
        except Exception as exc:
            self.lab_status_consulta.setText(f"Erro na consulta: {exc}")
            return
        self._popular_tabela_senha_corte(regs)
        self._ultimos_resultados_consulta = regs
        if ordem_txt:
            self.lab_status_consulta.setText(f"{len(regs)} registro(s) de Senha Corte para Ordem {ordem_txt}." if regs else f"Nenhum registro de Senha Corte para Ordem {ordem_txt}.")
        else:
            self.lab_status_consulta.setText(f"{len(regs)} registro(s) de Senha Corte." if regs else "Nenhum registro de Senha Corte no período.")

    def _popular_tabela_senha_corte(self, regs: list[dict]) -> None:
        headers = ["ID", "Data Ordem", "Ordem", "Carga", "Valor", "Tipo", "Observação", "Usuário", "Criado em", "ITEM", "Excluir"]
        self.tab_resultados.setSortingEnabled(False)
        self.tab_resultados.clearContents()
        self.tab_resultados.setRowCount(len(regs))
        self.tab_resultados.setColumnCount(len(headers))
        self.tab_resultados.setHorizontalHeaderLabels(headers)
        for row, r in enumerate(regs):
            # ID
            it_id = QTableWidgetItem()
            try:
                it_id.setData(Qt.ItemDataRole.DisplayRole, int(r.get("id")))
            except Exception:
                it_id.setData(Qt.ItemDataRole.DisplayRole, 0)
            self.tab_resultados.setItem(row, 0, it_id)
            # Data Ordem
            self.tab_resultados.setItem(row, 1, QTableWidgetItem(r.get("data_ordem", "")))
            # Ordem
            it_ordem = QTableWidgetItem()
            try:
                it_ordem.setData(Qt.ItemDataRole.DisplayRole, int(r.get("ordem")))
            except Exception:
                it_ordem.setData(Qt.ItemDataRole.DisplayRole, 0)
            self.tab_resultados.setItem(row, 2, it_ordem)
            # Carga
            it_carga = QTableWidgetItem()
            try:
                it_carga.setData(Qt.ItemDataRole.DisplayRole, int(r.get("carga")))
            except Exception:
                it_carga.setData(Qt.ItemDataRole.DisplayRole, 0)
            self.tab_resultados.setItem(row, 3, it_carga)
            # Valor (texto já formatado/normalizado do DB)
            self.tab_resultados.setItem(row, 4, QTableWidgetItem(r.get("valor", "")))
            # Tipo e Observação
            self.tab_resultados.setItem(row, 5, QTableWidgetItem(r.get("tipo_tratativa", "")))
            self.tab_resultados.setItem(row, 6, QTableWidgetItem(r.get("observacao", "")))
            # Usuário / Criado em (sempre em caixa alta)
            usuario_upper = (r.get("usuario") or "").upper()
            self.tab_resultados.setItem(row, 7, QTableWidgetItem(usuario_upper))
            self.tab_resultados.setItem(row, 8, QTableWidgetItem(r.get("created_at", "")))
            # ITEM
            from PySide6.QtWidgets import QPushButton
            btn = QPushButton(f"{r.get('item_count', 0)} item(ns)")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda _, sid=r.get("id"): self._mostrar_itens_de_senha_corte(sid))
            self.tab_resultados.setCellWidget(row, 9, btn)
            # Excluir
            from database import obter_tipo_usuario_atual, obter_usuario_atual
            tipo = obter_tipo_usuario_atual()
            usuario = obter_usuario_atual()
            can = (tipo == "ADMINISTRADOR" or (usuario and usuario == r.get("usuario")))
            btn_excluir = QPushButton("X")
            btn_excluir.setToolTip("Excluir este registro de Senha Corte")
            btn_excluir.setObjectName("BtnExcluir")
            btn_excluir.setStyleSheet(
                "QPushButton{background-color:#8B0000;color:#ffffff;border:1px solid #5a0000;padding:2px 8px;border-radius:4px;}"
                "QPushButton:hover{background-color:#a40000;}"
                "QPushButton:pressed{background-color:#7a0000;}"
            )
            btn_excluir.setEnabled(can)
            if can:
                btn_excluir.clicked.connect(lambda _, sid=r.get("id"): self._confirmar_e_excluir_senha_corte(sid))
            self.tab_resultados.setCellWidget(row, 10, btn_excluir)
        self.tab_resultados.setSortingEnabled(True)

    def _mostrar_itens_de_senha_corte(self, senha_id: int) -> None:
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QHBoxLayout
        from database import listar_itens_de_senha_corte
        try:
            itens = listar_itens_de_senha_corte(senha_id)
        except Exception as exc:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Senha Corte", f"Falha ao listar itens: {exc}")
            return
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Itens da Senha Corte #{senha_id}")
        v = QVBoxLayout(dlg)
        tab = QTableWidget(0, 3)
        tab.setHorizontalHeaderLabels(["Código", "Quantidade", "Tipo"])
        hdr = tab.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        try:
            hdr.setStyleSheet(
                "QHeaderView::section {background-color: #1976D2; color: #ffffff; font-weight: 600; padding: 6px 4px; border: none;}"
            )
        except Exception:
            pass
        tab.setAlternatingRowColors(True)
        tab.setRowCount(len(itens))
        for row, it in enumerate(itens):
            # Código
            it_cod = QTableWidgetItem()
            try:
                it_cod.setData(Qt.ItemDataRole.DisplayRole, int(it.get("codigo", 0)))
            except Exception:
                it_cod.setData(Qt.ItemDataRole.DisplayRole, 0)
            tab.setItem(row, 0, it_cod)
            # Quantidade
            it_qtd = QTableWidgetItem()
            try:
                it_qtd.setData(Qt.ItemDataRole.DisplayRole, int(it.get("quantidade", 0)))
            except Exception:
                it_qtd.setData(Qt.ItemDataRole.DisplayRole, 0)
            tab.setItem(row, 1, it_qtd)
            # Tipo
            tab.setItem(row, 2, QTableWidgetItem(it.get("tipo_tratativa", "")))
        v.addWidget(tab)
        row_btn = QHBoxLayout()
        row_btn.addStretch(1)
        btn_fechar = QPushButton("Fechar")
        btn_fechar.clicked.connect(dlg.accept)
        row_btn.addWidget(btn_fechar)
        v.addLayout(row_btn)
        dlg.resize(600, 300)
        dlg.exec()

    def _confirmar_e_excluir_senha_corte(self, senha_id: int) -> None:
        from PySide6.QtWidgets import QMessageBox
        resp = QMessageBox.question(self, "Confirmar exclusão", f"Deseja realmente excluir a Senha Corte #{senha_id}?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if resp != QMessageBox.StandardButton.Yes:
            return
        try:
            from database import excluir_senha_corte
            ok = excluir_senha_corte(senha_id)
        except Exception as exc:
            QMessageBox.critical(self, "Erro", f"Falha ao excluir: {exc}")
            return
        if ok:
            QMessageBox.information(self, "Excluído", "Registro de Senha Corte excluído com sucesso.")
            self._executar_consulta_senha_corte()
        else:
            QMessageBox.warning(self, "Aviso", "Sem permissão ou registro não encontrado.")

    def _executar_consulta_epis(self) -> None:
        di = self.ed_epi_data_ini.date().toString("yyyy-MM-dd") if hasattr(self, "ed_epi_data_ini") else None
        df = self.ed_epi_data_fim.date().toString("yyyy-MM-dd") if hasattr(self, "ed_epi_data_fim") else None
        mat = self.ed_epi_matricula.text().strip() if hasattr(self, "ed_epi_matricula") else ""
        try:
            if mat:
                from database import consultar_epis_por_matricula
                regs = consultar_epis_por_matricula(matricula=int(mat), data_ini=di, data_fim=df)
            else:
                from database import consultar_epis_por_periodo
                regs = consultar_epis_por_periodo(data_ini=di, data_fim=df)
        except ValueError as exc:
            self.lab_status_consulta.setText(str(exc))
            return
        except Exception as exc:
            self.lab_status_consulta.setText(f"Erro na consulta: {exc}")
            return
        self._popular_tabela_epis(regs)
        self._ultimos_resultados_consulta = regs
        self.lab_status_consulta.setText(f"{len(regs)} registro(s) de EPIs." if regs else "Nenhum registro de EPIs no filtro.")

    def _popular_tabela_epis(self, regs: list[dict]) -> None:
        headers = ["ID", "Data", "Matrícula", "Setor", "Turno", "Primeira", "Responsável", "Observação", "Usuário", "Criado em", "ITEM", "Excluir"]
        self.tab_resultados.setSortingEnabled(False)
        self.tab_resultados.clearContents()
        self.tab_resultados.setRowCount(len(regs))
        self.tab_resultados.setColumnCount(len(headers))
        self.tab_resultados.setHorizontalHeaderLabels(headers)
        for row, r in enumerate(regs):
            # Dados principais
            def _set_int(col, val):
                it = QTableWidgetItem()
                try:
                    it.setData(Qt.ItemDataRole.DisplayRole, int(val))
                except Exception:
                    it.setData(Qt.ItemDataRole.DisplayRole, 0)
                self.tab_resultados.setItem(row, col, it)
            _set_int(0, r.get("id"))
            self.tab_resultados.setItem(row, 1, QTableWidgetItem(r.get("data_ref", "")))
            _set_int(2, r.get("matricula"))
            self.tab_resultados.setItem(row, 3, QTableWidgetItem(r.get("setor", "")))
            self.tab_resultados.setItem(row, 4, QTableWidgetItem(r.get("turno", "")))
            self.tab_resultados.setItem(row, 5, QTableWidgetItem(r.get("primeira", "")))
            _set_int(6, r.get("responsavel"))
            self.tab_resultados.setItem(row, 7, QTableWidgetItem(r.get("observacao", "")))
            self.tab_resultados.setItem(row, 8, QTableWidgetItem(r.get("usuario") or ""))
            self.tab_resultados.setItem(row, 9, QTableWidgetItem(r.get("created_at", "")))
            # Coluna ITEM: mostra contagem e botão para detalhes
            from PySide6.QtWidgets import QPushButton
            btn = QPushButton(f"{r.get('item_count', 0)} item(ns)")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda _, eid=r.get("id"): self._mostrar_itens_de_epi(eid))
            self.tab_resultados.setCellWidget(row, 10, btn)
            # Coluna Excluir (X) com mesma aparência das demais exclusões
            from database import obter_tipo_usuario_atual, obter_usuario_atual
            tipo = obter_tipo_usuario_atual()
            usuario = obter_usuario_atual()
            can = (tipo == "ADMINISTRADOR" or (usuario and usuario == r.get("usuario")))
            btn_excluir = QPushButton("X")
            btn_excluir.setToolTip("Excluir este registro de EPIs")
            btn_excluir.setObjectName("BtnExcluir")
            btn_excluir.setStyleSheet(
                "QPushButton{background-color:#8B0000;color:#ffffff;border:1px solid #5a0000;padding:2px 8px;border-radius:4px;}"
                "QPushButton:hover{background-color:#a40000;}"
                "QPushButton:pressed{background-color:#7a0000;}"
            )
            btn_excluir.setEnabled(can)
            if can:
                btn_excluir.clicked.connect(lambda _, eid=r.get("id"): self._confirmar_e_excluir_epi(eid))
            self.tab_resultados.setCellWidget(row, 11, btn_excluir)
        self.tab_resultados.setSortingEnabled(True)

    def _confirmar_e_excluir_epi(self, epi_id: int) -> None:
        from PySide6.QtWidgets import QMessageBox
        resp = QMessageBox.question(self, "Confirmar exclusão", f"Deseja realmente excluir o registro de EPIs #{epi_id}?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if resp != QMessageBox.StandardButton.Yes:
            return
        try:
            from database import excluir_epi
            ok = excluir_epi(epi_id)
        except Exception as exc:
            QMessageBox.critical(self, "Erro", f"Falha ao excluir: {exc}")
            return
        if ok:
            QMessageBox.information(self, "Excluído", "Registro de EPIs excluído com sucesso.")
            # Recarrega a consulta atual de EPIs
            self._executar_consulta_epis()
        else:
            QMessageBox.warning(self, "Aviso", "Sem permissão ou registro não encontrado.")

    def _mostrar_itens_de_epi(self, epi_id: int) -> None:
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QHBoxLayout
        from PySide6.QtCore import Qt
        from database import listar_itens_de_epi
        try:
            itens = listar_itens_de_epi(epi_id)
        except Exception as exc:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "EPIs", f"Falha ao listar itens: {exc}")
            return
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Itens do EPI #{epi_id}")
        v = QVBoxLayout(dlg)
        tab = QTableWidget(0, 6)
        tab.setHorizontalHeaderLabels(["Código", "Produto", "Quantidade", "UON", "Valor Unit", "Valor Total"])
        # Estilo do cabeçalho: fundo azul e texto branco (padrão da consulta)
        hdr = tab.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        try:
            hdr.setStyleSheet(
                "QHeaderView::section {"
                "background-color: #1976D2;"  # azul padrão
                "color: #ffffff;"
                "font-weight: 600;"
                "padding: 6px 4px;"
                "border: none;"
                "}"
            )
        except Exception:
            pass
        tab.setAlternatingRowColors(True)
        tab.setRowCount(len(itens))
        for row, it in enumerate(itens):
            tab.setItem(row, 0, QTableWidgetItem(it.get("codigo", "")))
            tab.setItem(row, 1, QTableWidgetItem(it.get("produto", "")))
            q_it = QTableWidgetItem()
            try:
                q_it.setData(Qt.ItemDataRole.DisplayRole, int(it.get("quantidade", 0)))
            except Exception:
                q_it.setData(Qt.ItemDataRole.DisplayRole, 0)
            tab.setItem(row, 2, q_it)
            tab.setItem(row, 3, QTableWidgetItem(it.get("uon", "")))
            # Formatar valores monetários como BRL e alinhar à direita
            vu_txt = self._format_brl(it.get("valor_unit"))
            vt_txt = self._format_brl(it.get("valor_total"))
            vu_item = QTableWidgetItem(vu_txt)
            vu_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            vt_item = QTableWidgetItem(vt_txt)
            vt_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            tab.setItem(row, 4, vu_item)
            tab.setItem(row, 5, vt_item)
        v.addWidget(tab)
        row_btn = QHBoxLayout()
        row_btn.addStretch(1)
        btn_fechar = QPushButton("Fechar")
        btn_fechar.clicked.connect(dlg.accept)
        row_btn.addWidget(btn_fechar)
        v.addLayout(row_btn)
        dlg.resize(800, 400)
        dlg.exec()

    def _format_brl(self, valor) -> str:
        """Formata um valor (str/num) como BRL: R$ X.XXX,XX. Valores vazios viram R$ 0,00."""
        try:
            if valor is None or valor == "":
                v = 0.0
            else:
                s = str(valor).strip()
                # aceita "10", "10.00", "10,00"
                s = s.replace(".", "").replace(",", ".") if s.count(",") == 1 and s.count(".") > 1 else s.replace(",", ".")
                v = float(s)
            txt = f"{v:,.2f}"  # 1,234.56
            txt = txt.replace(",", "_").replace(".", ",").replace("_", ".")
            return f"R$ {txt}"
        except Exception:
            return f"R$ {valor}"

    def _executar_consulta_almox(self) -> None:
        ini = self.ed_alm_data_ini.date().toString("yyyy-MM-dd")
        fim = self.ed_alm_data_fim.date().toString("yyyy-MM-dd")
        turno_vis = self.cb_turno_alm.currentText()
        turno = None if turno_vis == "Todos" else turno_vis
        try:
            from database import consultar_almoxafire
            regs = consultar_almoxafire(turno=turno, data_ini=ini, data_fim=fim)
        except ValueError as exc:
            self.lab_status_consulta.setText(str(exc))
            return
        except Exception as exc:
            self.lab_status_consulta.setText(f"Erro na consulta: {exc}")
            return
        self._popular_tabela_almox(regs)
        self._ultimos_resultados_consulta = regs
        self.lab_status_consulta.setText(f"{len(regs)} registro(s) encontrados no Almoxarifado." if regs else "Nenhum registro encontrado.")

    def _popular_tabela_almox(self, regs: list[dict]) -> None:
        headers = ["id", "setor", "turno", "matricula", "responsavel", "insumo", "quantidade", "observacao", "usuario", "created_at", "editar", "excluir"]
        vis_headers = ["ID", "Setor", "Turno", "Matrícula", "Responsável", "Insumo", "Quantidade", "Observação", "Usuário", "Criado em", "Editar", "Excluir"]
        self.tab_resultados.setSortingEnabled(False)
        self.tab_resultados.clearContents()
        self.tab_resultados.setRowCount(len(regs))
        self.tab_resultados.setColumnCount(len(headers))
        self.tab_resultados.setHorizontalHeaderLabels(vis_headers)
        for row, r in enumerate(regs):
            # campos de dados
            dados_keys = ["id", "setor", "turno", "matricula", "responsavel", "insumo", "quantidade", "observacao", "usuario", "created_at"]
            for col, k in enumerate(dados_keys):
                val = r.get(k, "")
                if k in {"id", "matricula", "quantidade"}:
                    item = QTableWidgetItem()
                    try:
                        item.setData(Qt.ItemDataRole.DisplayRole, int(val))
                    except Exception:
                        item.setData(Qt.ItemDataRole.DisplayRole, 0)
                else:
                    item = QTableWidgetItem(val or "")
                self.tab_resultados.setItem(row, col, item)
            # ações
            from database import obter_tipo_usuario_atual, obter_usuario_atual
            tipo = obter_tipo_usuario_atual()
            usuario = obter_usuario_atual()
            can = (tipo == "ADMINISTRADOR" or (usuario and usuario == r.get("usuario")))
            btn_editar = QPushButton("Editar")
            btn_editar.setEnabled(can)
            if can:
                btn_editar.clicked.connect(lambda _, rid=r.get("id"), row=row: self._abrir_edicao_almox(rid, row))
            self.tab_resultados.setCellWidget(row, 10, btn_editar)

            btn_excluir = QPushButton("X")
            btn_excluir.setToolTip("Excluir este registro")
            btn_excluir.setObjectName("BtnExcluir")
            btn_excluir.setStyleSheet(
                "QPushButton{background-color:#8B0000;color:#ffffff;border:1px solid #5a0000;padding:2px 8px;border-radius:4px;}"
                "QPushButton:hover{background-color:#a40000;}"
                "QPushButton:pressed{background-color:#7a0000;}"
            )
            btn_excluir.setEnabled(can)
            if can:
                btn_excluir.clicked.connect(lambda _, rid=r.get("id"): self._confirmar_e_excluir_almox(rid))
            self.tab_resultados.setCellWidget(row, 11, btn_excluir)
        self.tab_resultados.setSortingEnabled(True)

    def _abrir_edicao_almox(self, almox_id: int, row: int) -> None:
        from PySide6.QtWidgets import QDialog, QFormLayout, QLineEdit, QComboBox, QTextEdit, QDialogButtonBox, QMessageBox
        from PySide6.QtGui import QIntValidator
        # Lê valores atuais da tabela
        def _txt(col):
            it = self.tab_resultados.item(row, col)
            return it.text() if it else ""
        atual = {
            "setor": _txt(1),
            "turno": _txt(2),
            "matricula": _txt(3),
            "responsavel": _txt(4),
            "insumo": _txt(5),
            "quantidade": _txt(6),
            "observacao": _txt(7),
        }
        class EditAlmoxDialog(QDialog):
            def __init__(self, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Editar Almoxarifado")
                lay = QFormLayout(self)
                self.cb_setor = QComboBox()
                setores = list(SETORES_GLOBAIS)
                self.cb_setor.addItems(setores)
                if atual["setor"] in setores:
                    self.cb_setor.setCurrentText(atual["setor"])
                self.cb_turno = QComboBox()
                self.cb_turno.addItems(["1° Turno", "2° Turno"])
                if atual["turno"] in ["1° Turno", "2° Turno"]:
                    self.cb_turno.setCurrentText(atual["turno"])
                self.ed_matricula = QLineEdit(atual["matricula"])
                self.ed_matricula.setValidator(QIntValidator(1, 99_999_999, self))
                self.ed_responsavel = QLineEdit(atual["responsavel"])
                self.cb_insumo = QComboBox()
                self.cb_insumo.addItems(["Fita", "Copo", "Caneta"])
                if atual["insumo"] in ["Fita", "Copo", "Caneta"]:
                    self.cb_insumo.setCurrentText(atual["insumo"])
                self.ed_qtd = QLineEdit(atual["quantidade"])
                self.ed_qtd.setValidator(QIntValidator(1, 1_000_000, self))
                self.ed_obs = QTextEdit()
                self.ed_obs.setAcceptRichText(False)
                self.ed_obs.setPlainText(atual["observacao"]) 
                lay.addRow("Setor:", self.cb_setor)
                lay.addRow("Turno:", self.cb_turno)
                lay.addRow("Matrícula:", self.ed_matricula)
                lay.addRow("Responsável:", self.ed_responsavel)
                lay.addRow("Insumo:", self.cb_insumo)
                lay.addRow("Quantidade:", self.ed_qtd)
                lay.addRow("Observação:", self.ed_obs)
                btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                btns.accepted.connect(self.accept)
                btns.rejected.connect(self.reject)
                lay.addRow(btns)
            def get_values(self):
                return {
                    "setor": self.cb_setor.currentText(),
                    "turno": self.cb_turno.currentText(),
                    "matricula": int(self.ed_matricula.text()),
                    "responsavel": self.ed_responsavel.text().strip(),
                    "insumo": self.cb_insumo.currentText(),
                    "quantidade": int(self.ed_qtd.text()),
                    "observacao": self.ed_obs.toPlainText().strip(),
                }
        dlg = EditAlmoxDialog(self)
        if not dlg.exec():
            return
        novos = dlg.get_values()
        try:
            from database import editar_almoxafire
            ok = editar_almoxafire(almox_id, **novos)
        except Exception as exc:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Erro", f"Falha ao editar: {exc}")
            return
        if ok:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Sucesso", "Registro atualizado.")
            self._executar_consulta_almox()
        else:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "Aviso", "Sem permissão ou registro não encontrado.")

    def _confirmar_e_excluir_almox(self, almox_id: int) -> None:
        from PySide6.QtWidgets import QMessageBox
        resp = QMessageBox.question(self, "Confirmar exclusão", f"Deseja realmente excluir o registro #{almox_id}?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if resp != QMessageBox.StandardButton.Yes:
            return
        try:
            from database import excluir_almoxafire
            ok = excluir_almoxafire(almox_id)
        except Exception as exc:
            QMessageBox.critical(self, "Erro", f"Falha ao excluir: {exc}")
            return
        if ok:
            QMessageBox.information(self, "Excluído", "Registro excluído com sucesso.")
            self._executar_consulta_almox()
        else:
            QMessageBox.warning(self, "Aviso", "Sem permissão ou registro não encontrado.")

    # -------- Popular tabelas -------- #
    def _popular_tabela_consolidado(self, regs: list[dict]) -> None:
        headers = [
            "data_ref",
            "arm.",
            "Descrição Filial",
            "R$ Estoque",
            "Qtde Item Mix",
            "Qtde Item com Estoque",
            "Qtde Item sem Estoque",
            "R$ Bloq. Total",
            "R$ Bloq. no ESTOQUE",
            "R$ Bloq. em Negoc.",
            "R$ Bloq. SALDO",
            "% Item com Estoque",
        ]
        self.tab_resultados.setSortingEnabled(False)
        self.tab_resultados.clear()
        self.tab_resultados.setColumnCount(len(headers))
        self.tab_resultados.setHorizontalHeaderLabels(headers)
        self.tab_resultados.clearContents()
        self.tab_resultados.setRowCount(len(regs))
        for row, r in enumerate(regs):
            for col, h in enumerate(headers):
                self.tab_resultados.setItem(row, col, QTableWidgetItem("" if r.get(h) is None else str(r.get(h))))
        self.tab_resultados.setSortingEnabled(True)

    def _popular_tabela_consultas(self, regs: list[dict]) -> None:
        was_sorting = self.tab_resultados.isSortingEnabled()
        self.tab_resultados.setSortingEnabled(False)
        row_count = len(regs)
        self.tab_resultados.clearContents()
        self.tab_resultados.setRowCount(row_count)
        for row, r in enumerate(regs):
            # ID
            item_id = QTableWidgetItem()
            try:
                item_id.setData(Qt.ItemDataRole.DisplayRole, int(r.get("id")))
            except Exception:
                item_id.setData(Qt.ItemDataRole.DisplayRole, 0)
            self.tab_resultados.setItem(row, 0, item_id)
            # Item
            item_cod = QTableWidgetItem()
            try:
                item_cod.setData(Qt.ItemDataRole.DisplayRole, int(r.get("item")))
            except Exception:
                item_cod.setData(Qt.ItemDataRole.DisplayRole, 0)
            self.tab_resultados.setItem(row, 1, item_cod)
            # Quantidade
            item_qtd = QTableWidgetItem()
            try:
                item_qtd.setData(Qt.ItemDataRole.DisplayRole, int(r.get("quantidade")))
            except Exception:
                item_qtd.setData(Qt.ItemDataRole.DisplayRole, 0)
            self.tab_resultados.setItem(row, 2, item_qtd)
            # Motivo
            self.tab_resultados.setItem(row, 3, QTableWidgetItem(r.get("motivo", "")))
            # Setor
            self.tab_resultados.setItem(row, 4, QTableWidgetItem(r.get("setor_responsavel", "")))
            # Matrícula
            self.tab_resultados.setItem(row, 5, QTableWidgetItem(str(r.get("matricula", ""))))
            # Data da movimentação
            self.tab_resultados.setItem(row, 6, QTableWidgetItem(r.get("data_mov", "")))
            # Usuário
            self.tab_resultados.setItem(row, 7, QTableWidgetItem(r.get("usuario") or ""))
            # Criado em
            self.tab_resultados.setItem(row, 8, QTableWidgetItem(r.get("created_at", "")))
            # Editar / Excluir
            btn_editar = QPushButton("Editar")
            btn_editar.setToolTip("Editar quantidade/setor")
            from database import obter_tipo_usuario_atual, obter_usuario_atual
            tipo = obter_tipo_usuario_atual()
            usuario = obter_usuario_atual()
            if tipo == "ADMINISTRADOR" or (usuario and usuario == r.get("usuario")):
                btn_editar.setEnabled(True)
                btn_editar.clicked.connect(lambda _, rid=r.get("id"), qtd=r.get("quantidade"), setor=r.get("setor_responsavel"): self._abrir_edicao_registro(rid, qtd, setor))
            else:
                btn_editar.setEnabled(False)
            self.tab_resultados.setCellWidget(row, 9, btn_editar)

            btn_excluir = QPushButton("X")
            btn_excluir.setToolTip("Excluir este registro")
            btn_excluir.setObjectName("BtnExcluir")
            btn_excluir.setStyleSheet(
                "QPushButton{background-color:#8B0000;color:#ffffff;border:1px solid #5a0000;padding:2px 8px;border-radius:4px;}"
                "QPushButton:hover{background-color:#a40000;}"
                "QPushButton:pressed{background-color:#7a0000;}"
            )
            if tipo == "ADMINISTRADOR" or (usuario and usuario == r.get("usuario")):
                btn_excluir.setEnabled(True)
                btn_excluir.clicked.connect(lambda _, rid=r.get("id"), item=r.get("item"), usr=r.get("usuario"): self._confirmar_e_excluir_registro(rid, item, usr))
            else:
                btn_excluir.setEnabled(False)
            self.tab_resultados.setCellWidget(row, 10, btn_excluir)
        self.tab_resultados.setSortingEnabled(was_sorting)

    def _popular_tabela_monitoramento(self, regs: list[dict]) -> None:
        was_sorting = self.tab_resultados.isSortingEnabled()
        self.tab_resultados.setSortingEnabled(False)
        self.tab_resultados.clearContents()
        self.tab_resultados.setRowCount(len(regs))
        for row, r in enumerate(regs):
            item_id = QTableWidgetItem()
            try:
                item_id.setData(Qt.ItemDataRole.DisplayRole, int(r.get("id")))
            except Exception:
                item_id.setData(Qt.ItemDataRole.DisplayRole, 0)
            self.tab_resultados.setItem(row, 0, item_id)
            self.tab_resultados.setItem(row, 1, QTableWidgetItem(r.get("onda", "")))
            self.tab_resultados.setItem(row, 2, QTableWidgetItem(r.get("carga", "")))
            self.tab_resultados.setItem(row, 3, QTableWidgetItem(r.get("container", "")))
            self.tab_resultados.setItem(row, 4, QTableWidgetItem(r.get("responsavel", "")))
            self.tab_resultados.setItem(row, 5, QTableWidgetItem(r.get("setor", "")))
            self.tab_resultados.setItem(row, 6, QTableWidgetItem(r.get("observacao", "")))
            self.tab_resultados.setItem(row, 7, QTableWidgetItem(r.get("usuario") or ""))
            self.tab_resultados.setItem(row, 8, QTableWidgetItem(r.get("created_at", "")))
            from database import obter_tipo_usuario_atual, obter_usuario_atual
            tipo = obter_tipo_usuario_atual()
            usuario = obter_usuario_atual()
            btn_editar = QPushButton("Editar")
            btn_editar.setToolTip("Editar responsável/setor/observação")
            if tipo == "ADMINISTRADOR" or (usuario and usuario == r.get("usuario")):
                btn_editar.setEnabled(True)
                btn_editar.clicked.connect(lambda _, rid=r.get("id"), resp=r.get("responsavel"), setor=r.get("setor"), obs=r.get("observacao"): self._abrir_edicao_monitoramento(rid, resp, setor, obs))
            else:
                btn_editar.setEnabled(False)
            self.tab_resultados.setCellWidget(row, 9, btn_editar)

            btn_excluir = QPushButton("X")
            btn_excluir.setToolTip("Excluir este registro")
            btn_excluir.setObjectName("BtnExcluir")
            btn_excluir.setStyleSheet(
                "QPushButton{background-color:#8B0000;color:#ffffff;border:1px solid #5a0000;padding:2px 8px;border-radius:4px;}"
                "QPushButton:hover{background-color:#a40000;}"
                "QPushButton:pressed{background-color:#7a0000;}"
            )
            if tipo == "ADMINISTRADOR" or (usuario and usuario == r.get("usuario")):
                btn_excluir.setEnabled(True)
                btn_excluir.clicked.connect(lambda _, rid=r.get("id"), ond=r.get("onda"), usr=r.get("usuario"): self._confirmar_e_excluir_monitoramento(rid, ond, usr))
            else:
                btn_excluir.setEnabled(False)
            self.tab_resultados.setCellWidget(row, 10, btn_excluir)
        self.tab_resultados.setSortingEnabled(was_sorting)

    # -------- Ações -------- #
    def _confirmar_e_excluir_registro(self, registro_id: int, item_cod: int, usuario_reg: Optional[str]) -> None:
        from PySide6.QtWidgets import QMessageBox
        alvo = f"registro #{registro_id} (item {item_cod})"
        if usuario_reg:
            alvo += f" de {usuario_reg}"
        resp = QMessageBox.question(self, "Confirmar exclusão", f"Deseja realmente excluir o {alvo}?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if resp != QMessageBox.StandardButton.Yes:
            return
        try:
            from database import excluir_registro
            ok = excluir_registro(registro_id)
        except Exception as exc:
            QMessageBox.critical(self, "Erro", f"Falha ao excluir: {exc}")
            return
        if ok:
            QMessageBox.information(self, "Excluído", "Registro excluído com sucesso.")
            self._executar_consulta_item()
        else:
            QMessageBox.warning(self, "Aviso", "Sem permissão ou registro não encontrado.")

    def _abrir_edicao_registro(self, registro_id, quantidade_atual, setor_atual):
        from PySide6.QtWidgets import QDialog, QFormLayout, QLineEdit, QComboBox, QDialogButtonBox, QMessageBox
        from PySide6.QtGui import QIntValidator
        class EditDialog(QDialog):
            def __init__(self, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Editar Registro")
                lay = QFormLayout(self)
                self.ed_quantidade = QLineEdit(str(quantidade_atual))
                self.ed_quantidade.setValidator(QIntValidator(1, 1_000_000, self))
                self.cb_setor = QComboBox()
                setores = list(SETORES_GLOBAIS)
                self.cb_setor.addItems(setores)
                if setor_atual in setores:
                    self.cb_setor.setCurrentText(setor_atual)
                lay.addRow("Quantidade:", self.ed_quantidade)
                lay.addRow("Setor:", self.cb_setor)
                btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                btns.accepted.connect(self.accept)
                btns.rejected.connect(self.reject)
                lay.addRow(btns)
            def get_values(self):
                return int(self.ed_quantidade.text()), self.cb_setor.currentText()
        dlg = EditDialog(self)
        if dlg.exec():
            nova_qtd, novo_setor = dlg.get_values()
            try:
                from database import editar_registro
                ok = editar_registro(registro_id, quantidade=nova_qtd, setor_responsavel=novo_setor)
            except Exception as exc:
                QMessageBox.critical(self, "Erro", f"Falha ao editar: {exc}")
                return
            if ok:
                QMessageBox.information(self, "Sucesso", "Registro atualizado.")
                self._executar_consulta_item()
            else:
                QMessageBox.warning(self, "Aviso", "Sem permissão ou registro não encontrado.")

    def _abrir_edicao_monitoramento(self, registro_id, responsavel_atual, setor_atual, observacao_atual):
        from PySide6.QtWidgets import QDialog, QFormLayout, QLineEdit, QComboBox, QTextEdit, QDialogButtonBox, QMessageBox
        class EditMonDialog(QDialog):
            def __init__(self, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Editar Reimpressão")
                lay = QFormLayout(self)
                self.ed_responsavel = QLineEdit(str(responsavel_atual or ""))
                self.cb_setor = QComboBox()
                setores = list(SETORES_GLOBAIS)
                self.cb_setor.addItems(setores)
                if setor_atual in setores:
                    self.cb_setor.setCurrentText(setor_atual)
                self.ed_obs = QTextEdit()
                self.ed_obs.setAcceptRichText(False)
                self.ed_obs.setPlainText(str(observacao_atual or ""))
                lay.addRow("Responsável:", self.ed_responsavel)
                lay.addRow("Setor:", self.cb_setor)
                lay.addRow("Observação:", self.ed_obs)
                btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                btns.accepted.connect(self.accept)
                btns.rejected.connect(self.reject)
                lay.addRow(btns)
            def get_values(self):
                return self.ed_responsavel.text().strip(), self.cb_setor.currentText(), self.ed_obs.toPlainText().strip()
        dlg = EditMonDialog(self)
        if dlg.exec():
            novo_resp, novo_setor, nova_obs = dlg.get_values()
            try:
                from database import editar_monitoramento
                ok = editar_monitoramento(registro_id, responsavel=novo_resp, setor=novo_setor, observacao=nova_obs)
            except Exception as exc:
                QMessageBox.critical(self, "Erro", f"Falha ao editar: {exc}")
                return
            if ok:
                QMessageBox.information(self, "Sucesso", "Registro atualizado.")
                self._executar_consulta_monitoramento()
            else:
                QMessageBox.warning(self, "Aviso", "Sem permissão ou registro não encontrado.")

    def _confirmar_e_excluir_monitoramento(self, registro_id: int, onda: str, usuario_reg: Optional[str]) -> None:
        from PySide6.QtWidgets import QMessageBox
        alvo = f"registro #{registro_id} (onda {onda})"
        if usuario_reg:
            alvo += f" de {usuario_reg}"
        resp = QMessageBox.question(self, "Confirmar exclusão", f"Deseja realmente excluir o {alvo}?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if resp != QMessageBox.StandardButton.Yes:
            return
        try:
            from database import excluir_monitoramento
            ok = excluir_monitoramento(registro_id)
        except Exception as exc:
            QMessageBox.critical(self, "Erro", f"Falha ao excluir: {exc}")
            return
        if ok:
            QMessageBox.information(self, "Excluído", "Registro excluído com sucesso.")
            self._executar_consulta_monitoramento()
        else:
            QMessageBox.warning(self, "Aviso", "Sem permissão ou registro não encontrado.")

    def _exportar_consulta(self) -> None:
        regs = getattr(self, "_ultimos_resultados_consulta", [])
        if not regs:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Exportar", "Nenhum resultado para exportar. Faça uma consulta primeiro.")
            return
        fmt_dlg = ConsultaExportDialog(self)
        if not fmt_dlg.exec():
            return
        formato = fmt_dlg.get_formato().lower()
        from PySide6.QtWidgets import QFileDialog
        from datetime import datetime
        padrao_nome = f"consultas_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{formato}"
        if formato == "csv":
            filtro = "CSV (*.csv)"
        elif formato == "xlsx":
            filtro = "Excel (*.xlsx)"
        else:
            filtro = "Texto (*.txt)"
        caminho, _ = QFileDialog.getSaveFileName(self, "Salvar resultados", padrao_nome, filtro)
        if not caminho:
            return
        try:
            fonte = self.cb_fonte.currentText() if hasattr(self, "cb_fonte") else "Bloqueado"
            if fonte == "Consolidado":
                headers = [
                    "data_ref",
                    "arm.",
                    "Descrição Filial",
                    "R$ Estoque",
                    "Qtde Item Mix",
                    "Qtde Item com Estoque",
                    "Qtde Item sem Estoque",
                    "R$ Bloq. Total",
                    "R$ Bloq. no ESTOQUE",
                    "R$ Bloq. em Negoc.",
                    "R$ Bloq. SALDO",
                    "% Item com Estoque",
                ]
                if formato == "csv":
                    import csv
                    with open(caminho, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f, delimiter=';')
                        w.writerow(headers)
                        for r in regs:
                            w.writerow([r.get(h, "") if r.get(h) is not None else "" for h in headers])
                elif formato == "xlsx":
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.append(headers)
                    for r in regs:
                        ws.append([r.get(h, "") if r.get(h) is not None else "" for h in headers])
                    wb.save(caminho)
                else:
                    with open(caminho, "w", encoding="utf-8") as f:
                        for r in regs:
                            f.write("\t".join([str(r.get(h, "") or "") for h in headers]) + "\n")
            elif fonte == "Reimpressão":
                headers = ["id", "onda", "carga", "container", "responsavel", "setor", "observacao", "usuario", "created_at"]
                if formato == "csv":
                    import csv
                    with open(caminho, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f, delimiter=';')
                        w.writerow(["ID", "Onda", "Carga", "Container", "Responsável", "Setor", "Observação", "Usuário", "Criado em"])
                        for r in regs:
                            w.writerow([r.get(k, "") or "" for k in headers])
                elif formato == "xlsx":
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["ID", "Onda", "Carga", "Container", "Responsável", "Setor", "Observação", "Usuário", "Criado em"])
                    for r in regs:
                        ws.append([r.get(k, "") or "" for k in headers])
                    wb.save(caminho)
                else:
                    with open(caminho, "w", encoding="utf-8") as f:
                        for r in regs:
                            f.write("\t".join([str(r.get(k, "") or "") for k in headers]) + "\n")
            elif fonte == "Almoxarifado":
                headers = ["id", "setor", "turno", "matricula", "responsavel", "insumo", "quantidade", "observacao", "usuario", "created_at"]
                vis_headers = ["ID", "Setor", "Turno", "Matrícula", "Responsável", "Insumo", "Quantidade", "Observação", "Usuário", "Criado em"]
                if formato == "csv":
                    import csv
                    with open(caminho, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f, delimiter=';')
                        w.writerow(vis_headers)
                        for r in regs:
                            w.writerow([r.get(k, "") or "" for k in headers])
                elif formato == "xlsx":
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.append(vis_headers)
                    for r in regs:
                        ws.append([r.get(k, "") or "" for k in headers])
                    wb.save(caminho)
                else:
                    with open(caminho, "w", encoding="utf-8") as f:
                        for r in regs:
                            f.write("\t".join([str(r.get(k, "") or "") for k in headers]) + "\n")
            elif fonte == "EPIs":
                headers = ["id", "data_ref", "matricula", "setor", "turno", "primeira", "responsavel", "observacao", "usuario", "created_at", "item_count"]
                vis_headers = ["ID", "Data", "Matrícula", "Setor", "Turno", "Primeira", "Responsável", "Observação", "Usuário", "Criado em", "Itens"]
                if formato == "csv":
                    import csv
                    with open(caminho, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f, delimiter=';')
                        w.writerow(vis_headers)
                        for r in regs:
                            w.writerow([r.get(k, "") or "" for k in headers])
                elif formato == "xlsx":
                    from openpyxl import Workbook
                    from database import listar_itens_de_epi
                    wb = Workbook()
                    # Planilha 1: EPIS (cabeçalhos)
                    ws1 = wb.active
                    ws1.title = "EPIS"
                    ws1.append(vis_headers)
                    for r in regs:
                        ws1.append([r.get(k, "") or "" for k in headers])
                    # Planilha 2: EPIS_ITENS (itens de todos os EPIs)
                    ws2 = wb.create_sheet("EPIS_ITENS")
                    itens_headers = ["epi_id", "id", "codigo", "produto", "quantidade", "uon", "valor_unit", "valor_total"]
                    ws2.append(itens_headers)
                    for r in regs:
                        eid = r.get("id")
                        if not eid:
                            continue
                        try:
                            itens = listar_itens_de_epi(eid)
                        except Exception:
                            itens = []
                        for it in itens:
                            ws2.append([
                                eid,
                                it.get("id", ""),
                                it.get("codigo", ""),
                                it.get("produto", ""),
                                it.get("quantidade", ""),
                                it.get("uon", ""),
                                it.get("valor_unit", ""),
                                it.get("valor_total", ""),
                            ])
                    wb.save(caminho)
                else:
                    with open(caminho, "w", encoding="utf-8") as f:
                        for r in regs:
                            f.write("\t".join([str(r.get(k, "") or "") for k in headers]) + "\n")
            elif fonte == "Senha Corte":
                headers = ["id", "data_ordem", "ordem", "carga", "valor", "tipo_tratativa", "observacao", "usuario", "created_at", "item_count"]
                vis_headers = ["ID", "Data Ordem", "Ordem", "Carga", "Valor", "Tipo", "Observação", "Usuário", "Criado em", "Itens"]
                if formato == "csv":
                    import csv
                    with open(caminho, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f, delimiter=';')
                        w.writerow(vis_headers)
                        for r in regs:
                            w.writerow([r.get(k, "") or "" for k in headers])
                elif formato == "xlsx":
                    from openpyxl import Workbook
                    from database import listar_itens_de_senha_corte
                    wb = Workbook()
                    ws1 = wb.active
                    ws1.title = "SENHA_CORTE"
                    ws1.append(vis_headers)
                    for r in regs:
                        ws1.append([r.get(k, "") or "" for k in headers])
                    ws2 = wb.create_sheet("SENHA_CORTE_ITENS")
                    itens_headers = ["senha_id", "id", "codigo", "quantidade", "tipo_tratativa"]
                    ws2.append(itens_headers)
                    for r in regs:
                        sid = r.get("id")
                        if not sid:
                            continue
                        try:
                            itens = listar_itens_de_senha_corte(sid)
                        except Exception:
                            itens = []
                        for it in itens:
                            ws2.append([
                                sid,
                                it.get("id", ""),
                                it.get("codigo", ""),
                                it.get("quantidade", ""),
                                it.get("tipo_tratativa", ""),
                            ])
                    wb.save(caminho)
                else:
                    with open(caminho, "w", encoding="utf-8") as f:
                        for r in regs:
                            f.write("\t".join([str(r.get(k, "") or "") for k in headers]) + "\n")
            else:
                # Bloqueado (inclui Data da movimentação)
                headers = ["id", "item", "quantidade", "motivo", "setor_responsavel", "matricula", "data_mov", "usuario", "created_at"]
                vis_headers = ["ID", "Item", "Quantidade", "Motivo", "Setor", "Matrícula", "Data Mov.", "Usuário", "Criado em"]
                if formato == "csv":
                    import csv
                    with open(caminho, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f, delimiter=';')
                        w.writerow(vis_headers)
                        for r in regs:
                            w.writerow([r.get(k, "") or "" for k in headers])
                elif formato == "xlsx":
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.append(vis_headers)
                    for r in regs:
                        ws.append([r.get(k, "") or "" for k in headers])
                    wb.save(caminho)
                else:
                    with open(caminho, "w", encoding="utf-8") as f:
                        for r in regs:
                            f.write("\t".join([str(r.get(k, "") or "") for k in headers]) + "\n")
        except Exception as exc:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Erro", f"Falha ao exportar: {exc}")
            return
        from PySide6.QtWidgets import QMessageBox
        QMessageBox.information(self, "Exportar", f"Exportado {len(regs)} registro(s).")
