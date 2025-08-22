from __future__ import annotations

from typing import Optional

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QFileDialog,
    QMessageBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractScrollArea,
)


class ConsolidadoPage(QWidget):
    """Página Consolidado com importação de XLSX e visualização em tabela.

    - Mantém padrão de tabela da página Consultas (objectName, estilos, comportamento)
    - Mensagens de erro exibidas em label vermelho dentro da própria página
    - Leitura com pandas (preferencial) e fallback com openpyxl
    """

    REQUIRED_COLS = [
        "arm.",
        "Descrição Filial",
        "R$ Estoque",
        "Qtde Item Mix",
        "Qtde Item com Estoque",
        "Qtde Item sem Estoque",
        "R$ Bloq. Total",
        "R$ Bloq. no ESTOQUE",
        "R$ Bloq. em Negoc.",
        "R$ Bloq. SALDO",
        "% Item com Estoque",
    ]

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setObjectName("PaginaConsolidado")
        self._arquivo: Optional[str] = None
        self._build()

    def _build(self) -> None:
        lay = QVBoxLayout(self)
        lay.setContentsMargins(32, 24, 32, 24)
        lay.setSpacing(16)

        # Cabeçalho simples (reutiliza estilo do Bloqueado, se aplicado globalmente)
        header = QLabel("CONSOLIDADO")
        header.setObjectName("TituloBloqueado")
        header.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        lay.addWidget(header)

        # Seleção de arquivo
        row = QHBoxLayout()
        self.lab_arquivo = QLabel("Nenhum arquivo selecionado")
        self.lab_arquivo.setObjectName("StatusLabel")
        btn_sel = QPushButton("Selecionar XLSX")
        btn_sel.clicked.connect(self._selecionar_arquivo)
        row.addWidget(self.lab_arquivo, 1)
        row.addWidget(btn_sel)
        wrap = QWidget()
        wrap.setLayout(row)
        lay.addWidget(wrap)

        # Ações
        actions = QHBoxLayout()
        actions.addStretch(1)
        btn_enviar = QPushButton("Enviar")
        btn_enviar.setObjectName("Primary")
        btn_enviar.clicked.connect(self._enviar)
        actions.addWidget(btn_enviar)
        # Botão de salvar no banco (aparece somente após sucesso)
        self.btn_salvar = QPushButton("SALVAR NO BANCO DE DADOS")
        self.btn_salvar.setVisible(False)
        self.btn_salvar.clicked.connect(self._salvar_no_banco)
        actions.addWidget(self.btn_salvar)
        lay.addLayout(actions)

        # Status de erro (vermelho)
        self.lab_status = QLabel("")
        self.lab_status.setObjectName("ConsolidadoStatusLabel")
        self.lab_status.setStyleSheet("#ConsolidadoStatusLabel { color: #b00020; padding:4px 2px; }")
        self.lab_status.setVisible(False)
        lay.addWidget(self.lab_status)

        # Tabela (padrão igual à Consultas)
        self.tab_consolidado = QTableWidget(0, 0)
        self.tab_consolidado.setObjectName("TabelaConsultas")
        hdr = self.tab_consolidado.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        hdr.setStretchLastSection(True)
        self.tab_consolidado.setSizeAdjustPolicy(QAbstractScrollArea.SizeAdjustPolicy.AdjustToContents)
        self.tab_consolidado.setAlternatingRowColors(True)
        self.tab_consolidado.setSortingEnabled(True)
        lay.addWidget(self.tab_consolidado, 1)

        # Aplica o QSS da página Consultas (visual consistente)
        try:
            from style import QSS_CONSULTAS_PAGE
            self.setStyleSheet(self.styleSheet() + QSS_CONSULTAS_PAGE)
        except Exception:
            pass

    # --------- slots --------- #
    def _selecionar_arquivo(self) -> None:
        caminho, _ = QFileDialog.getOpenFileName(self, "Selecionar arquivo XLSX", "", "Excel (*.xlsx)")
        if caminho:
            self._arquivo = caminho
            self.lab_arquivo.setText(caminho)
            self._clear_status()

    def _enviar(self) -> None:
        if not self._arquivo:
            self._set_error("Selecione um arquivo XLSX primeiro.")
            return
        # Tenta com pandas primeiro
        try:
            import pandas as pd  # type: ignore
            df = pd.read_excel(self._arquivo, header=1)
            # Verifica colunas obrigatórias
            cols_encontradas = set(str(c) for c in df.columns.tolist())
            if set(self.REQUIRED_COLS).issubset(cols_encontradas):
                # Filtro e colunas padrão (mantém comportamento anterior)
                try:
                    consolidado_importado = df[df["arm."] == 1001][self.REQUIRED_COLS]
                except Exception:
                    consolidado_importado = df[self.REQUIRED_COLS]
                self._mostrar_dataframe(consolidado_importado)
                self._set_success("Consolidado recebido com sucesso")
                self.btn_salvar.setVisible(True)
                return
            else:
                # Mostra conteúdo bruto para diagnóstico e informa não ser consolidado
                self._mostrar_dataframe(df)
                self._set_error("Arquivo não é Consolidado de estoque")
                self.btn_salvar.setVisible(False)
                return
        except Exception as exc_pd:
            self._set_error(f"Falha ao ler com pandas: {exc_pd}")
            self.btn_salvar.setVisible(False)
        # Fallback: openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self._arquivo, read_only=True, data_only=True)
            if wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
                # Extrai cabeçalho da primeira linha para validar colunas
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers = [str(h) if h is not None else "" for h in rows[0]]
                    if set(self.REQUIRED_COLS).issubset(set(headers)):
                        self._mostrar_worksheet(ws)
                        self._set_success("Consolidado recebido com sucesso")
                        self.btn_salvar.setVisible(True)
                        wb.close()
                        return
                    else:
                        self._mostrar_worksheet(ws)
                        self._set_error("Arquivo não é Consolidado de estoque")
                        self.btn_salvar.setVisible(False)
                        wb.close()
                        return
            wb.close()
            # Se chegou aqui, não conseguiu validar/mostrar
            self._set_error("Falha ao validar o arquivo (sem abas)")
            self.btn_salvar.setVisible(False)
        except Exception as exc_ox:
            self._set_error(f"Falha ao abrir com openpyxl: {exc_ox}")
            self.btn_salvar.setVisible(False)

    # --------- helpers de status --------- #
    def _set_error(self, mensagem: str) -> None:
        # volta o estilo para vermelho sempre que houver erro
        self.lab_status.setStyleSheet("#ConsolidadoStatusLabel { color: #b00020; padding:4px 2px; }")
        self.lab_status.setText(mensagem)
        self.lab_status.setVisible(True)
        # esconde o botão de salvar em caso de erro
        self.btn_salvar.setVisible(False)

    def _clear_status(self) -> None:
        self.lab_status.clear()
        self.lab_status.setVisible(False)

    def _set_success(self, mensagem: str) -> None:
        # Verde compatível com temas claros/escuros
        self.lab_status.setStyleSheet("#ConsolidadoStatusLabel { color: #1e8e3e; padding:4px 2px; }")
        self.lab_status.setText(mensagem)
        self.lab_status.setVisible(True)
        # Restaura estilo vermelho default para próximos erros
        # (feito ao chamar _set_error novamente)

    # --------- salvar no banco --------- #
    def _salvar_no_banco(self) -> None:
        """Salva o consolidado exibido na tabela na tabela 'CONSOLIDADO'.

        Verifica se já existe consolidado para a data de hoje e pergunta se deve substituir.
        """
        try:
            # Extrai dados atuais da tabela para JSON
            headers = [self.tab_consolidado.horizontalHeaderItem(i).text() for i in range(self.tab_consolidado.columnCount())]
            linhas = []
            for r in range(self.tab_consolidado.rowCount()):
                linha = {}
                for c, h in enumerate(headers):
                    item = self.tab_consolidado.item(r, c)
                    linha[h] = item.text() if item else ""
                linhas.append(linha)
            # Data de referência = hoje (sem hora)
            from datetime import date
            from database import (
                existe_consolidado_linhas_na_data,
                salvar_consolidado_linhas,
                substituir_consolidado_linhas_por_data,
            )
            hoje = date.today()
            existe_linhas = existe_consolidado_linhas_na_data(data_ref=hoje)
            if existe_linhas:
                resp = QMessageBox.question(
                    self,
                    "Confirmar substituição",
                    "Já existe Consolidado para hoje. Deseja substituir?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if resp != QMessageBox.StandardButton.Yes:
                    return
                substituir_consolidado_linhas_por_data(data_ref=hoje, linhas=linhas)
            else:
                salvar_consolidado_linhas(data_ref=hoje, linhas=linhas)
            QMessageBox.information(self, "Consolidado", "Consolidado salvo no banco.")
        except Exception as exc:
            self._set_error(f"Falha ao salvar no banco: {exc}")

    # --------- helpers de exibição --------- #
    def _mostrar_dataframe(self, df) -> None:
        """Exibe um pandas.DataFrame na tabela padrão de Consultas."""
        try:
            import pandas as pd  # type: ignore
            if not isinstance(df, pd.DataFrame):
                return
            df = df.fillna("")
            headers = [str(c) for c in df.columns.tolist()]
            self.tab_consolidado.setSortingEnabled(False)
            self.tab_consolidado.clear()
            self.tab_consolidado.setColumnCount(len(headers))
            self.tab_consolidado.setHorizontalHeaderLabels(headers)
            self.tab_consolidado.setRowCount(len(df))
            for r_idx, (_, row) in enumerate(df.iterrows()):
                for c_idx, col in enumerate(headers):
                    val = row[col]
                    item = QTableWidgetItem(str(val))
                    self.tab_consolidado.setItem(r_idx, c_idx, item)
            self.tab_consolidado.setSortingEnabled(True)
        except Exception as exc:
            self._set_error(f"Falha ao exibir dados: {exc}")

    def _mostrar_worksheet(self, ws) -> None:
        """Exibe uma worksheet do openpyxl na tabela (cabeçalho na primeira linha)."""
        try:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                self.tab_consolidado.clear()
                self.tab_consolidado.setRowCount(0)
                self.tab_consolidado.setColumnCount(0)
                return
            headers = [str(h) if h is not None else "" for h in rows[0]]
            data = rows[1:]
            self.tab_consolidado.setSortingEnabled(False)
            self.tab_consolidado.clear()
            self.tab_consolidado.setColumnCount(len(headers))
            self.tab_consolidado.setHorizontalHeaderLabels(headers)
            self.tab_consolidado.setRowCount(len(data))
            for r_idx, r in enumerate(data):
                for c_idx, val in enumerate(r):
                    item = QTableWidgetItem("" if val is None else str(val))
                    self.tab_consolidado.setItem(r_idx, c_idx, item)
            self.tab_consolidado.setSortingEnabled(True)
        except Exception as exc:
            self._set_error(f"Falha ao exibir dados (openpyxl): {exc}")
